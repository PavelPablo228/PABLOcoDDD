import csv
from openpyxl import load_workbook

def parse_xlsx_by_surname(filename, surname_column=1, output_file='result.csv', encoding='utf-8'):
    workbook = load_workbook(filename, read_only=True, data_only=True)
    sheet = workbook.active

    surnames = set()
    for row in sheet.iter_rows(values_only=True, min_row=2):
        if surname_column < 1 or surname_column > len(row):
            raise ValueError("new number of culumn")

        surname = row[surname_column - 1]
        if surname is not None: 
            try:
                surnames.add(str(surname).lower().strip())
            except ValueError:
                print(f"error change: {surname}")

    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        for surname in surnames:
            writer.writerow([surname])

    return surnames

filename = r'C:\Users\Pavel Pablo\Desktop\КнигаААА.xlsx'
output_file = 'names5.csv'
unique_surnames = parse_xlsx_by_surname(filename, output_file=output_file)
print(f"Результаты сохранены в файл {output_file}")
