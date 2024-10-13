import datetime
import csv
from openpyxl import load_workbook

def parse_xlsx_by_surname(filename, surname_column=1, output_file='result.csv', encoding='utf-8'):
 
  workbook = load_workbook(filename, read_only=True, data_only=True)
  sheet = workbook.active

  surnames = set()
  for row in sheet.iter_rows(values_only=True, min_row=2):
    surname = row[surname_column - 1]
    if isinstance(surname, datetime.datetime):
      surname = surname.strftime('%Y-%m-%d %H:%M:%S')
    surnames.add(surname.lower().strip())

  with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile)
    for surname in surnames:
      writer.writerow([surname])

filename = r'C:\Users\Pavel Pablo\Desktop\КнигаААА.xlsx'
output_file = 'names1.csv'
unique_surnames = parse_xlsx_by_surname(filename, output_file=output_file)
print(f"Результаты сохранены в файл {output_file}")
